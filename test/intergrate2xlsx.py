import os
import natsort
import openpyxl
import pandas as pd 

def copy(file, index):
    game = openpyxl.load_workbook(os.path.join('./result/',file))
    gameSheet = game.active
    name = 'G'+str(index)
    sheet = wb.create_sheet(name)

    for i in range(1, 75):
        for j in range(1, 26):
           sheet.cell(row=i,column=j).value = gameSheet.cell(row=i,column=j).value


files = os.listdir('./result')
files = natsort.natsorted(files)
files = [file for file in files if file.endswith(".xlsx")]
wb = openpyxl.Workbook()

for i, file in enumerate(files):
  copy(file, i+1)

wb.remove(wb['Sheet'])
wb.save("intergrate.xlsx")

a = pd.read_excel("intergrate.xlsx", header=None) 
a.to_csv("intergrate.csv")